# export_manager.py
# WDI Visit Analytics Engine
# Generates styled Excel workbooks for export.

import io
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ─────────────────────────────────────────────
# BRAND COLOURS (hex without #)
# ─────────────────────────────────────────────
PRIMARY   = "1F4E79"
SECONDARY = "2E75B6"
ACCENT    = "70AD47"
HEADER_FG = "FFFFFF"
ALT_ROW   = "EBF3FB"
BORDER_COLOR = "B8CCE4"


def _side() -> Side:
    return Side(border_style="thin", color=BORDER_COLOR)


def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(color: str = PRIMARY) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=ALT_ROW)


def _header_font() -> Font:
    return Font(bold=True, color=HEADER_FG, name="Calibri", size=11)


def _normal_font(bold: bool = False) -> Font:
    return Font(bold=bold, name="Calibri", size=10)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center", wrap_text=True)


def _auto_width(ws, min_w: int = 12, max_w: int = 45):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 4))


def _write_df_to_sheet(ws, df: pd.DataFrame, header_color: str = PRIMARY, start_row: int = 1):
    """Write a DataFrame to an openpyxl worksheet with styling."""
    # Write header
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.fill   = _header_fill(header_color)
        cell.font   = _header_font()
        cell.border = _border()
        cell.alignment = _center()

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        fill = _alt_fill() if (row_idx - start_row) % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.border    = _border()
            cell.font      = _normal_font()
            cell.alignment = _right_align() if _is_arabic_col(df.columns[col_idx - 1]) else _center()

    ws.row_dimensions[start_row].height = 25
    _auto_width(ws)


def _is_arabic_col(col_name: str) -> bool:
    arabic_cols = {
        "Customer Name", "Visit Notes", "Governorate", "District",
        "Sales Rep Name", "Classification Reason", "Matched Keywords",
        "Status History", "Latest Status", "Display Status",
    }
    return col_name in arabic_cols


def _add_title_row(ws, title: str, num_cols: int, color: str = PRIMARY):
    """Merge first row as a report title."""
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill      = _header_fill(color)
    cell.font      = Font(bold=True, color=HEADER_FG, name="Calibri", size=13)
    cell.alignment = _center()
    ws.row_dimensions[1].height = 30


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


# ═══════════════════════════════════════════════════════════════════
# PUBLIC EXPORT FUNCTIONS
# ═══════════════════════════════════════════════════════════════════

def export_customer_summary(journey_df: pd.DataFrame) -> bytes:
    """
    Export: Customer Summary.xlsx
    Columns: Customer Name, Latest Status, Visit Count,
             First Visit Date, Last Visit Date, Days Since Last Visit,
             Governorate, District, Latest Confidence, Status History
    """
    export_cols = [
        "Customer Name", "Latest Status", "Visit Count",
        "First Visit Date", "Last Visit Date", "Days Since Last Visit",
        "Governorate", "District", "Latest Confidence", "Status History",
    ]
    df = journey_df.copy()
    # Drop internal columns
    for col in ["_journey"]:
        if col in df.columns:
            df = df.drop(columns=[col])

    # Keep only export columns that exist
    present = [c for c in export_cols if c in df.columns]
    df = df[present]

    # Format dates
    for dc in ["First Visit Date", "Last Visit Date"]:
        if dc in df.columns:
            df[dc] = pd.to_datetime(df[dc], errors="coerce").dt.strftime("%Y-%m-%d")

    if "Latest Confidence" in df.columns:
        df["Latest Confidence"] = df["Latest Confidence"].apply(lambda x: f"{x:.1f}%")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Summary"
    _write_df_to_sheet(ws, df, header_color=PRIMARY)
    _add_title_row(ws, f"WDI — Customer Summary  |  Generated: {_timestamp()}", len(df.columns))
    ws.sheet_view.rightToLeft = True  # RTL for Arabic

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sales_rep_kpi(rep_kpi_df: pd.DataFrame) -> bytes:
    """
    Export: Sales Rep KPI.xlsx
    """
    df = rep_kpi_df.copy()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Rep KPI"
    _write_df_to_sheet(ws, df, header_color=SECONDARY)
    _add_title_row(ws, f"WDI — Sales Rep KPI  |  Generated: {_timestamp()}", len(df.columns), color=SECONDARY)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_executive_dashboard(summary_dict: dict) -> bytes:
    """
    Export: Executive Dashboard.xlsx
    summary_dict keys: kpis (dict), monthly (df), status_dist (df), gov_dist (df), top_customers (df)
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: KPIs ──
    ws_kpi = wb.active
    ws_kpi.title = "KPIs"
    kpis = summary_dict.get("kpis", {})
    kpi_df = pd.DataFrame(
        [{"Metric": k, "Value": v} for k, v in kpis.items()]
    )
    _write_df_to_sheet(ws_kpi, kpi_df, header_color=PRIMARY)
    _add_title_row(ws_kpi, f"WDI — Executive KPIs  |  {_timestamp()}", 2)

    # ── Sheet 2: Monthly Trend ──
    monthly_df = summary_dict.get("monthly")
    if monthly_df is not None and not monthly_df.empty:
        ws_m = wb.create_sheet("Monthly Trend")
        _write_df_to_sheet(ws_m, monthly_df, header_color=SECONDARY)
        _add_title_row(ws_m, "Monthly Visit Trend", len(monthly_df.columns), color=SECONDARY)

    # ── Sheet 3: Status Distribution ──
    status_df = summary_dict.get("status_dist")
    if status_df is not None and not status_df.empty:
        ws_s = wb.create_sheet("Status Distribution")
        _write_df_to_sheet(ws_s, status_df, header_color=ACCENT)
        _add_title_row(ws_s, "Customer Status Distribution", len(status_df.columns), color=ACCENT)

    # ── Sheet 4: Governorate ──
    gov_df = summary_dict.get("gov_dist")
    if gov_df is not None and not gov_df.empty:
        ws_g = wb.create_sheet("Governorate Distribution")
        _write_df_to_sheet(ws_g, gov_df, header_color=PRIMARY)
        _add_title_row(ws_g, "Governorate Distribution", len(gov_df.columns))

    # ── Sheet 5: Top Customers ──
    top_df = summary_dict.get("top_customers")
    if top_df is not None and not top_df.empty:
        ws_t = wb.create_sheet("Top Customers")
        _write_df_to_sheet(ws_t, top_df, header_color=SECONDARY)
        _add_title_row(ws_t, "Top 20 Most Visited Customers", len(top_df.columns), color=SECONDARY)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_classification_results(classified_df: pd.DataFrame) -> bytes:
    """
    Export: Classification Results.xlsx
    """
    export_cols = [
        "Visit Date", "Customer Name", "Sales Rep Name",
        "Suggested Status", "Display Status", "Confidence Score",
        "Raw Score", "Matched Keywords", "Classification Reason",
        "Governorate", "District", "Visit Notes",
    ]
    df = classified_df.copy()

    # Format dates
    if "Visit Date" in df.columns:
        df["Visit Date"] = pd.to_datetime(df["Visit Date"], errors="coerce").dt.strftime("%Y-%m-%d")

    if "Confidence Score" in df.columns:
        df["Confidence Score"] = df["Confidence Score"].apply(lambda x: f"{x:.1f}%")

    present = [c for c in export_cols if c in df.columns]
    df = df[present]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classification Results"
    ws.sheet_view.rightToLeft = True

    _write_df_to_sheet(ws, df, header_color=PRIMARY)
    _add_title_row(ws, f"WDI — Classification Results  |  {_timestamp()}", len(df.columns))

    # Freeze panes below header title rows
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_followup_customers(journey_df: pd.DataFrame, days_threshold: int = 30) -> bytes:
    """
    Export customers who haven't been visited for >= days_threshold days.
    """
    df = journey_df.copy()
    for col in ["_journey"]:
        if col in df.columns:
            df = df.drop(columns=[col])

    df = df[df["Days Since Last Visit"].fillna(9999) >= days_threshold].copy()
    df = df.sort_values("Days Since Last Visit", ascending=False)

    for dc in ["First Visit Date", "Last Visit Date"]:
        if dc in df.columns:
            df[dc] = pd.to_datetime(df[dc], errors="coerce").dt.strftime("%Y-%m-%d")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"No Visit {days_threshold}+ Days"
    ws.sheet_view.rightToLeft = True
    _write_df_to_sheet(ws, df, header_color="C00000")
    _add_title_row(
        ws,
        f"WDI — Customers Not Visited {days_threshold}+ Days  |  {_timestamp()}",
        len(df.columns),
        color="C00000",
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
